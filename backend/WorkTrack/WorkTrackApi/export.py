from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import os
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from rest_framework.exceptions import ValidationError
from datetime import date, datetime, timedelta, time
from calendar import monthrange
from .models import Attendance, Employees, CalendarDay,PlannedShifts,TypeShift
class MonthlyRosterExporter:
    def __init__(self, year, month):
        self.year = year
        self.month = month
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = f"Rozpis {month}-{year}"
        
        # Konštanty a štýly
        self.thin = Side(style='thin')
        self.thick = Side(style='medium')
        self.bold_font = Font(bold=True)
        self.title_font = Font(bold=True, size=14)
        self.center_align = Alignment(horizontal='center', vertical='center')
        self.right_align = Alignment(horizontal='right', vertical='center')
        self.vertical_align = Alignment(horizontal='center', vertical='center', textRotation=90)
        self.gray_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # Mapovania
        self.SHIFT_TO_COLUMN_MAP = {
            23: 'Ič', 21: 'Do', 16: 'D', 13: 'D', 15: 'D', 14: 'D',
            19: 'Pp', 17: 'Pp', 18: 'Pp', 20: 'Ns', 
            8: 'SN', 9: 'SN', 11: 'SN', 12: 'SN', 10: 'SN', 22: 'Ič',
        }
        self.ID_TO_TEXT_MAP = {21: 'DO', 23: 'PN'}
        self.SUB_COLUMNS = ['SN', 'Do', 'Pp', 'Ič', 'D', 'Ns']
        self.NON_WORKING_IDS = [21, 23]

    def _get_shift_duration(self, s_start, s_end):
        if not (s_start and s_end): return 0.0
        dt_s = datetime.combine(date(2000, 1, 1), s_start)
        dt_e = datetime.combine(date(2000, 1, 1), s_end)
        if dt_e < dt_s: dt_e += timedelta(days=1)
        if s_end == time(0, 0) and s_start != s_end: dt_e += timedelta(days=1)
        return (dt_e - dt_s).total_seconds() / 3600.0

    def _calculate_overlap_seconds(self, start1, end1, start2, end2):
        latest_start = max(start1, start2)
        earliest_end = min(end1, end2)
        return max(0.0, (earliest_end - latest_start).total_seconds())

    def generate_response(self):
        """Hlavná metóda, ktorá spustí generovanie a vráti HTTP Response"""
        
        # 1. Príprava dát
        slovak_months = ["", "Január", "Február", "Marec", "Apríl", "Máj", "Jún", 
                         "Júl", "August", "September", "Október", "November", "December"]
        month_name = slovak_months[self.month]
        
        holidays_dates = set(val[0] for val in CalendarDay.get_slovak_holidays(self.year))
        days_in_month = monthrange(self.year, self.month)[1]
        start_date = date(self.year, self.month, 1)
        end_date = date(self.year, self.month, days_in_month)

        work_days_count = sum(1 for d in range(1, days_in_month + 1) 
                              if date(self.year, self.month, d).weekday() < 5 and date(self.year, self.month, d) not in holidays_dates)
        
        STANDARD_HOURS = 7.0
        work_fund = work_days_count * STANDARD_HOURS

        # 2. Načítanie zamestnancov (Filter)
        active_ids_in_month = PlannedShifts.objects.filter(
            date__range=(start_date, end_date), hidden=False
        ).values_list('user_id', flat=True).distinct()

        employees = Employees.objects.filter(id__in=active_ids_in_month).order_by('last_name')
        
        if not employees.exists():
            raise ValidationError(detail=f"Pre obdobie {self.month}/{self.year} neboli nájdené žiadne platné smeny na export.")
        
        num_emps = employees.count()
        last_col_idx = 2 + (num_emps * 6)

        # 3. Vykreslenie HLAVIČKY
        self.ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col_idx)
        self.ws.cell(row=1, column=1, value=f"Rozdeľovník služieb 5.SuS {month_name} {self.year}").font = self.title_font
        self.ws.cell(row=1, column=1).alignment = self.center_align

        self.ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
        self.ws.cell(row=2, column=1, value="Počet pracovných dní:").font = self.bold_font
        self.ws.cell(row=2, column=5, value=work_days_count).alignment = self.center_align

        mid_col = max(5, last_col_idx // 2)
        self.ws.cell(row=2, column=mid_col, value=month_name).font = self.bold_font
        self.ws.cell(row=2, column=mid_col).alignment = self.center_align

        self.ws.merge_cells(start_row=2, start_column=last_col_idx-3, end_row=2, end_column=last_col_idx-1)
        self.ws.cell(row=2, column=last_col_idx-3, value="Úväzok v hod.:").font = self.bold_font
        self.ws.cell(row=2, column=last_col_idx-3).alignment = self.right_align
        self.ws.cell(row=2, column=last_col_idx, value=work_fund).alignment = self.center_align

        # 4. Vykreslenie MENÁ A TABUĽKA
        NAME_ROW, SUB_HEADER_ROW, DATA_START_ROW = 4, 5, 6

        # Vertikálne nadpisy Dátum/Deň
        for col_idx, val in enumerate(["Dátum", "Deň"], start=1):
            self.ws.merge_cells(start_row=NAME_ROW, start_column=col_idx, end_row=SUB_HEADER_ROW, end_column=col_idx)
            c = self.ws.cell(row=NAME_ROW, column=col_idx, value=val)
            c.alignment = self.vertical_align
            c.font = Font(size=9, bold=True)
            # Rámčeky Dátum/Deň
            c.border = Border(top=self.thick, left=(self.thick if col_idx==1 else self.thin), 
                              right=(self.thick if col_idx==2 else self.thin))
            self.ws.cell(row=SUB_HEADER_ROW, column=col_idx).border = Border(bottom=self.thick, 
                              left=(self.thick if col_idx==1 else self.thin), 
                              right=(self.thick if col_idx==2 else self.thin))

        emp_col_start, current_col = {}, 3
        
        # Slučka cez zamestnancov (Hlavička)
        border_thick_box = Border(left=self.thick, right=self.thick, top=self.thick, bottom=self.thick)
        
        for emp in employees:
            self.ws.merge_cells(start_row=NAME_ROW, start_column=current_col, end_row=NAME_ROW, end_column=current_col + 5)
            c = self.ws.cell(row=NAME_ROW, column=current_col, value=f"{emp.last_name} {emp.first_name[0]}.")
            c.font = self.bold_font
            c.alignment = self.center_align
            for i in range(6): self.ws.cell(row=NAME_ROW, column=current_col + i).border = border_thick_box
            
            for i, col_name in enumerate(self.SUB_COLUMNS):
                actual_col = current_col + i
                c = self.ws.cell(row=SUB_HEADER_ROW, column=actual_col, value=col_name)
                c.alignment = self.center_align
                c.font = Font(size=8)
                self.ws.column_dimensions[get_column_letter(actual_col)].width = 4
                
                b_left = self.thick if i == 0 else self.thin
                b_right = self.thick if i == 5 else self.thin
                c.border = Border(left=b_left, right=b_right, top=self.thin, bottom=self.thick)

            emp_col_start[emp.id] = current_col
            current_col += 6

        # 5. DÁTA (Mriežka a Smeny)
        shifts = PlannedShifts.objects.filter(date__range=(start_date, end_date), hidden=False).select_related('type_shift')
        shift_data = {}
        for s in shifts:
            d_iso = s.date.isoformat()
            shift_data.setdefault(d_iso, {}).setdefault(s.user_id, []).append(s)

        slovak_days = ['Po', 'Ut', 'St', 'Št', 'Pi', 'So', 'Ne']
        for day in range(1, days_in_month + 1):
            row_idx = day + DATA_START_ROW - 1
            curr_date = date(self.year, self.month, day)
            d_iso = curr_date.isoformat()
            is_weekend = curr_date.weekday() >= 5
            
            # Prvé dva stĺpce
            self.ws.cell(row=row_idx, column=1, value=f"{day}.").border = Border(left=self.thick, right=self.thin, top=self.thin, bottom=self.thin)
            self.ws.cell(row=row_idx, column=2, value=slovak_days[curr_date.weekday()]).border = Border(left=self.thin, right=self.thick, top=self.thin, bottom=self.thin)
            if is_weekend:
                for c_idx in [1, 2]: self.ws.cell(row=row_idx, column=c_idx).fill = self.gray_fill

            # Mriežka zamestnancov
            for col in range(3, current_col):
                cell = self.ws.cell(row=row_idx, column=col)
                pos = (col - 3) % 6
                cell.border = Border(left=(self.thick if pos == 0 else self.thin), right=(self.thick if pos == 5 else self.thin), 
                                     top=self.thin, bottom=(self.thick if day == days_in_month else self.thin))
                if is_weekend: cell.fill = self.gray_fill

            # Vyplnenie smien
            if d_iso in shift_data:
                for user_id, user_shifts in shift_data[d_iso].items():
                    start_col = emp_col_start.get(user_id)
                    if not start_col: continue 
                    for shift in user_shifts:
                        if not shift.type_shift or (shift.type_shift.id == 20 and shift.custom_start == time(0, 0)): continue
                        target_col = self.SHIFT_TO_COLUMN_MAP.get(shift.type_shift.id) or (shift.type_shift.shortName if shift.type_shift.shortName in self.SUB_COLUMNS else None)
                        if not target_col: continue
                        final_col = start_col + self.SUB_COLUMNS.index(target_col)
                        
                        val = self.ID_TO_TEXT_MAP.get(shift.type_shift.id)
                        if val is None:
                            val = float(shift.type_shift.duration_time) if shift.type_shift.duration_time else self._get_shift_duration(shift.custom_start, shift.custom_end)
                            if val == 0: val = ""
                            elif val % 1 == 0: val = int(val)
                            else: val = round(val, 2)

                        cell = self.ws.cell(row=row_idx, column=final_col, value=val)
                        cell.alignment = self.center_align
                        # Obnova rámu po zápise
                        pos = (final_col - 3) % 6
                        cell.border = Border(left=(self.thick if pos == 0 else self.thin), right=(self.thick if pos == 5 else self.thin), 
                                             top=self.thin, bottom=(self.thick if day == days_in_month else self.thin))
                        if is_weekend: cell.fill = self.gray_fill

        # 6. SUMÁRNE VÝPOČTY
        SUMMARY_START_ROW = days_in_month + DATA_START_ROW + 1
        labels = ["Prenesené hodiny:", "Odpracované hodiny:", "+/- za mesiac:", "SPOLU za mesiac:", "Sviatky:", "Sobota:", "Nedeľa:", "Nočná smena:"]
        
        for i, label in enumerate(labels):
            row = SUMMARY_START_ROW + i
            self.ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            c = self.ws.cell(row=row, column=1, value=label)
            c.font, c.alignment = self.bold_font, self.right_align
            self.ws.cell(row=row, column=1).border = Border(left=self.thick, top=self.thin, bottom=self.thin)
            self.ws.cell(row=row, column=2).border = Border(right=self.thick, top=self.thin, bottom=self.thin)

        shifts_by_user = {}
        for s in shifts: shifts_by_user.setdefault(s.user_id, []).append(s)

        for emp in employees:
            start_col = emp_col_start.get(emp.id)
            user_shifts = shifts_by_user.get(emp.id, [])
            total_worked = holiday_hours = saturday_hours = sunday_hours = night_hours = 0.0
            transferred = float(emp.initial_hours_balance or 0.0)
            
            for s in user_shifts:
                duration = float(s.type_shift.duration_time) if s.type_shift and s.type_shift.duration_time else self._get_shift_duration(s.custom_start, s.custom_end)
                total_worked += duration
                if s.type_shift.id in self.NON_WORKING_IDS or not (s.custom_start and s.custom_end): continue
                
                dt_s = datetime.combine(s.date, s.custom_start)
                dt_e = datetime.combine(s.date, s.custom_end)
                if dt_e < dt_s: dt_e += timedelta(days=1)
                
                curr = dt_s
                while curr < dt_e:
                    seg_end = min(dt_e, datetime.combine(curr.date() + timedelta(days=1), time(0,0)))
                    seg_duration = (seg_end - curr).total_seconds() / 3600.0
                    if curr.weekday() == 5: saturday_hours += seg_duration
                    elif curr.weekday() == 6: sunday_hours += seg_duration
                    if curr.date() in holidays_dates: holiday_hours += seg_duration
                    curr = seg_end
                
                check_d, final_d = dt_s.date() - timedelta(days=1), dt_e.date()
                while check_d <= final_d:
                    night_hours += self._calculate_overlap_seconds(dt_s, dt_e, datetime.combine(check_d, time(22, 0)), datetime.combine(check_d + timedelta(days=1), time(6, 0))) / 3600.0
                    check_d += timedelta(days=1)

            diff = total_worked - work_fund
            vals = [transferred, total_worked, diff, transferred + diff, holiday_hours, saturday_hours, sunday_hours, night_hours]
            for i, v in enumerate(vals):
                row = SUMMARY_START_ROW + i
                self.ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col + 5)
                c = self.ws.cell(row=row, column=start_col, value=round(v, 2))
                c.alignment, c.font = self.center_align, Font(bold=(i==3))
                for k in range(6):
                    bc = self.ws.cell(row=row, column=start_col + k)
                    bc.border = Border(left=(self.thick if k==0 else self.thin), right=(self.thick if k==5 else self.thin), top=self.thin, bottom=self.thin)

        # 7. DYNAMICKÁ LEGENDA
        categories = [
            {"title": "sobota, nedeľa", "ids": [8, 9, 10, 11, 12]},
            {"title": "Dopoludnia", "ids": [13, 14, 15]},
            {"title": "Popoludní", "ids": [17, 18, 19]},
            {"title": "Nočná služba", "ids": [20]},
            {"title": "iná činnosť", "ids": [22]}
        ]

        LEGEND_START_ROW, curr_legend_col = SUMMARY_START_ROW + len(labels) + 2, 1
        BLOCK_WIDTH = 4

        for cat in categories:
            items = []
            for ts in TypeShift.objects.filter(id__in=cat["ids"]):
                h = str(ts.duration_time) if ts.duration_time else ""
                t_range = f"{ts.start_time.strftime('%H,%M')}-{ts.end_time.strftime('%H,%M')}" if ts.start_time and ts.end_time else ""
                if [h, t_range] not in items: items.append([h, t_range])
            
            # Nadpis
            self.ws.merge_cells(start_row=LEGEND_START_ROW, start_column=curr_legend_col, end_row=LEGEND_START_ROW, end_column=curr_legend_col + BLOCK_WIDTH - 1)
            tc = self.ws.cell(row=LEGEND_START_ROW, column=curr_legend_col, value=cat["title"])
            tc.font, tc.alignment = self.bold_font, self.center_align
            self.ws.cell(row=LEGEND_START_ROW, column=curr_legend_col).border = Border(top=self.thick, left=self.thick, bottom=self.thin)
            self.ws.cell(row=LEGEND_START_ROW, column=curr_legend_col + BLOCK_WIDTH - 1).border = Border(top=self.thick, right=self.thick, bottom=self.thin)
            for k in range(1, BLOCK_WIDTH - 1): self.ws.cell(row=LEGEND_START_ROW, column=curr_legend_col + k).border = Border(top=self.thick, bottom=self.thin)

            # Podnadpisy
            c_lbl_h = self.ws.cell(row=LEGEND_START_ROW + 1, column=curr_legend_col, value="hodiny")
            c_lbl_h.font, c_lbl_h.alignment = Font(size=8, bold=True), self.center_align
            c_lbl_h.border = Border(left=self.thick, right=self.thin, top=self.thin, bottom=self.thin)

            self.ws.merge_cells(start_row=LEGEND_START_ROW + 1, start_column=curr_legend_col + 1, end_row=LEGEND_START_ROW + 1, end_column=curr_legend_col + BLOCK_WIDTH - 1)
            c_lbl_t = self.ws.cell(row=LEGEND_START_ROW + 1, column=curr_legend_col + 1, value="v čase")
            c_lbl_t.font, c_lbl_t.alignment = Font(size=8, bold=True), self.center_align
            c_lbl_t.border = Border(left=self.thin, top=self.thin, bottom=self.thin)
            self.ws.cell(row=LEGEND_START_ROW + 1, column=curr_legend_col + BLOCK_WIDTH - 1).border = Border(right=self.thick, top=self.thin, bottom=self.thin)

            # Dáta
            for i in range(4):
                row_idx = LEGEND_START_ROW + 2 + i
                val_h, val_t = items[i] if i < len(items) else ("", "")
                
                ch = self.ws.cell(row=row_idx, column=curr_legend_col, value=val_h)
                ch.alignment = self.center_align
                
                self.ws.merge_cells(start_row=row_idx, start_column=curr_legend_col + 1, end_row=row_idx, end_column=curr_legend_col + BLOCK_WIDTH - 1)
                ct = self.ws.cell(row=row_idx, column=curr_legend_col + 1, value=val_t)
                ct.alignment = self.center_align

                b_bottom = self.thick if i == 3 else self.thin
                ch.border = Border(left=self.thick, right=self.thin, top=self.thin, bottom=b_bottom)
                ct.border = Border(left=self.thin, top=self.thin, bottom=b_bottom)
                self.ws.cell(row=row_idx, column=curr_legend_col + BLOCK_WIDTH - 1).border = Border(right=self.thick, top=self.thin, bottom=b_bottom)
                for k in range(2, BLOCK_WIDTH - 1): self.ws.cell(row=row_idx, column=curr_legend_col + k).border = Border(top=self.thin, bottom=b_bottom)

            curr_legend_col += BLOCK_WIDTH

        # 8. Uloženie a Response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=Rozpis_FULL_{self.month}_{self.year}.xlsx'
        self.wb.save(response)
        return response
    




# class AttendancePdfExporter:
#     def __init__(self, employee_id, year, month):
#         self.employee_id = employee_id
#         self.year = year
#         self.month = month
        
#         # Ošetrenie, ak zamestnanec neexistuje
#         try:
#             self.employee = Employees.objects.get(pk=employee_id)
#         except Employees.DoesNotExist:
#              self.employee = None

#         # --- REGISTRÁCIA FONTU (Arial) ---
#         # Skúsime nájsť Arial vo Windows priečinku fontov
#         # Ak nie sme na Windowse alebo font neexistuje, použijeme defaultnú Helveticu
        
#         self.custom_font = 'Helvetica'
#         self.custom_font_bold = 'Helvetica-Bold'
        
#         try:
#             # Cesta k fontom vo Windows
#             fonts_dir = os.path.join(os.environ.get('WINDIR', 'C:\\Windows'), 'Fonts')
#             arial_path = os.path.join(fonts_dir, 'arial.ttf')
#             arialbd_path = os.path.join(fonts_dir, 'arialbd.ttf')
            
#             # Overenie existencie súborov
#             if os.path.exists(arial_path) and os.path.exists(arialbd_path):
#                 # Registrácia fontov pod vlastnými názvami
#                 pdfmetrics.registerFont(TTFont('ArialCustom', arial_path))
#                 pdfmetrics.registerFont(TTFont('ArialBoldCustom', arialbd_path))
                
#                 # Nastavenie premenných na použitie nových fontov
#                 self.custom_font = 'ArialCustom'
#                 self.custom_font_bold = 'ArialBoldCustom'
#         except Exception:
#             # V prípade akejkoľvek chyby pri registrácii fontu ostaneme pri Helvetice
#             pass
        
#         # Nastavenie štýlov
#         self.styles = getSampleStyleSheet()
#         # Vytvoríme vlastný štýl pre normálny text s našim fontom
#         self.styles.add(ParagraphStyle(name='CustomNormal', parent=self.styles['Normal'], fontName=self.custom_font, fontSize=10))
#         # Vytvoríme vlastný štýl pre nadpis s našim tučným fontom
#         self.styles.add(ParagraphStyle(name='CustomTitle', parent=self.styles['Title'], fontName=self.custom_font_bold))
        
#         self.buffer = BytesIO()
#         self.doc = SimpleDocTemplate(
#             self.buffer, 
#             pagesize=landscape(A4),
#             rightMargin=10*mm, leftMargin=10*mm, 
#             topMargin=10*mm, bottomMargin=10*mm
#         )
#         self.elements = []
#         self.notes_list = [] # Inicializácia zoznamu poznámok

#     def _format_time(self, t):
#         if not t: return ""
#         return t.strftime("%H,%M")

#     def _format_duration(self, seconds):
#         if not seconds: return ""
#         hours = seconds / 3600.0
#         if hours % 1 == 0:
#             return f"{int(hours)}"
#         return f"{hours:.1f}".replace(".", ",")

#     def generate_response(self):
#         # 1. Zber Dát
#         days_in_month = monthrange(self.year, self.month)[1]
#         start_date = date(self.year, self.month, 1)
#         end_date = date(self.year, self.month, days_in_month)
        
#         attendances = Attendance.objects.filter(
#             user_id=self.employee_id,
#             date__range=(start_date, end_date)
#         ).select_related('type_shift').order_by('date')
        
#         att_by_date = {a.date: a for a in attendances}
#         holidays = set(val[0] for val in CalendarDay.get_slovak_holidays(self.year))

#         slovak_months = ["", "Január", "Február", "Marec", "Apríl", "Máj", "Jún", 
#                          "Júl", "August", "September", "Október", "November", "December"]

#         # 2. Hlavička dokumentu (Texty)
#         header_text = f"<b>Centrum pre deti a rodiny Poprad</b><br/>Výkaz pedagogických pracovníkov, pomocných vychovávateľov a ostatných pracovníkov za"
#         self.elements.append(Paragraph(header_text, self.styles['CustomTitle']))
#         self.elements.append(Spacer(1, 5*mm))

#         info_data = [
#             [f"Mesiac / Rok: {slovak_months[self.month]} {self.year}", 
#              f"Meno: {self.employee.last_name} {self.employee.first_name}" if self.employee else "", 
#              f"Osobné číslo: {self.employee.personal_number or ''}" if self.employee else ""]
#         ]
#         info_table = Table(info_data, colWidths=[80*mm, 120*mm, 60*mm])
#         info_table.setStyle(TableStyle([
#             ('FONTNAME', (0,0), (-1,-1), self.custom_font_bold), 
#             ('FONTSIZE', (0,0), (-1,-1), 10),
#             ('ALIGN', (0,0), (-1,-1), 'LEFT'),
#         ]))
#         self.elements.append(info_table)
#         self.elements.append(Spacer(1, 5*mm))

#         # 3. TABUĽKA - HLAVIČKA
#         table_data = []
        
#         # Prvý riadok hlavičky (Skupiny)
#         headers_row1 = [
#             "",              # Nad "Deň" (0)
#             "Dopoludnia",    # (1)
#             "",              # (2) - zlúčené s 1
#             "Odpoludnia",    # (3)
#             "",              # (4) - zlúčené s 3
#             "",              # Nad "Odp.hod." (5)
#             "Príplatky",     # (6)
#             "",              # (7) - zlúčené s 6
#             "",              # (8) - zlúčené s 6
#             "",              # (9) - zlúčené s 6
#             "",              # (10) - zlúčené s 6
#             "Neodpr. hod.",  # (11)
#             ""               # (12) - zlúčené s 11
#         ]
        
#         # Druhý riadok hlavičky (Konkrétne stĺpce)
#         headers_row2 = [
#             "Deň",           # 0
#             "Príchod",       # 1  (Dopoludnia)
#             "Odchod",        # 2  (Dopoludnia)
#             "Príchod",       # 3  (Odpoludnia)
#             "Odchod",        # 4  (Odpoludnia)
#             "Odp.hod.",      # 5
#             "Sob.",          # 6
#             "Ned.",          # 7
#             "Iná\nčinn.",    # 8
#             "Sviatok",       # 9
#             "Nočná",         # 10
#             "Dovol.",        # 11
#             "PN"             # 12
#         ]
        
#         table_data.append(headers_row1)
#         table_data.append(headers_row2)

#         total_worked = 0.0
#         sum_sob, sum_ned, sum_ina, sum_sviatok, sum_noc, sum_dov, sum_pn = 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0
        
#         # 4. Riadky dní
#         for day in range(1, days_in_month + 1):
#             curr_date = date(self.year, self.month, day)
#             att = att_by_date.get(curr_date)
#             row = [f"{day}."] + [""] * 12
            
#             if att:
#                 s_time = att.custom_start
#                 e_time = att.custom_end
#                 duration = 0.0
                
#                 # Výpočet trvania
#                 if att.type_shift and att.type_shift.duration_time:
#                      duration = float(att.type_shift.duration_time)
#                 elif s_time and e_time:
#                     dt_s = datetime.combine(curr_date, s_time)
#                     dt_e = datetime.combine(curr_date, e_time)
#                     if dt_e < dt_s: dt_e += timedelta(days=1)
#                     if e_time == time(0,0) and s_time != e_time: dt_e += timedelta(days=1)
#                     duration = (dt_e - dt_s).total_seconds() / 3600.0

#                 # --- HLAVNÁ ZMENA LOGIKY ---
#                 # Zápis časov: Dopoludnia vs Odpoludnia
#                 # index 1, 2 = Dopoludnia
#                 # index 3, 4 = Odpoludnia
                
#                 # Definujeme ID pre absencie, pri ktorých nepíšeme časy
#                 ABSENCE_IDS = [21, 23] 

#                 if att.type_shift and att.type_shift.id not in ABSENCE_IDS:
#                     if s_time and e_time:
#                         NOON = time(12, 0)
                        
#                         # 1. ZÁPIS PRÍCHODU (Start)
#                         # Ak začína pred 12:00 -> Dopoludnia (idx 1), inak Odpoludnia (idx 3)
#                         if s_time < NOON:
#                             row[1] = self._format_time(s_time)
#                         else:
#                             row[3] = self._format_time(s_time)

#                         # 2. ZÁPIS ODCHODU (End)
#                         # Logika: 
#                         # - Ak smena prechádza cez polnoc (e_time < s_time), končí určite "poobede/ráno ďalšieho dňa" -> Dávame do Odpoludnia (idx 4)
#                         # - Ak končí po 12:00 -> Odpoludnia (idx 4)
#                         # - Ak končí pred 12:00 -> Dopoludnia (idx 2)
                        
#                         is_overnight = e_time < s_time # Nočná smena
                        
#                         if is_overnight or e_time >= NOON:
#                             row[4] = self._format_time(e_time) # Odchod Odpoludnia
#                         else:
#                             row[2] = self._format_time(e_time) # Odchod Dopoludnia

#                     # Zápis odpracovaných hodín
#                     if duration > 0:
#                         row[5] = self._format_duration(duration * 3600)
#                         total_worked += duration

#                 # --- PRÍPLATKY (Stĺpce 6-12) ---
#                 if curr_date.weekday() == 5: # Sobota
#                     row[6] = self._format_duration(duration * 3600)
#                     sum_sob += duration
#                 elif curr_date.weekday() == 6: # Nedeľa
#                     row[7] = self._format_duration(duration * 3600)
#                     sum_ned += duration
                
#                 if att.type_shift and att.type_shift.id == 22: # Iná činnosť
#                     row[8] = self._format_duration(duration * 3600)
#                     sum_ina += duration
                
#                 if curr_date in holidays and att.type_shift.id not in ABSENCE_IDS: # Sviatok
#                      row[9] = self._format_duration(duration * 3600)
#                      sum_sviatok += duration

#                 if att.type_shift and att.type_shift.id == 20: # Nočná
#                      row[10] = self._format_duration(duration * 3600)
#                      sum_noc += duration
                
#                 if att.type_shift and att.type_shift.id == 21: # Dovolenka
#                     row[11] = "D"
#                     sum_dov += 7.5 
                
#                 if att.type_shift and att.type_shift.id == 23: # PN
#                     row[12] = "PN"
#                     sum_pn += 7.5
                
#                 if att.note:
#                     self.notes_list.append(f"{day}. {self.month}. - {att.note}")

#             table_data.append(row)

#         # Riadok Spolu
#         row_total = ["Spolu", "", "", "", "", 
#                      self._format_duration(total_worked * 3600),
#                      self._format_duration(sum_sob * 3600),
#                      self._format_duration(sum_ned * 3600),
#                      self._format_duration(sum_ina * 3600),
#                      self._format_duration(sum_sviatok * 3600),
#                      self._format_duration(sum_noc * 3600),
#                      self._format_duration(sum_dov * 3600) if sum_dov else "",
#                      self._format_duration(sum_pn * 3600) if sum_pn else ""]
#         table_data.append(row_total)

#         # Vykreslenie Tabuľky
#         col_widths = [10*mm] + [18*mm]*4 + [15*mm]*8
#         t = Table(table_data, colWidths=col_widths, repeatRows=2)
        
#         # Štýlovanie tabuľky
#         t.setStyle(TableStyle([
#             ('GRID', (0,0), (-1,-1), 0.5, colors.black),
#             ('FONTNAME', (0,0), (-1,-1), self.custom_font),
#             ('FONTSIZE', (0,0), (-1,-1), 8),
#             ('ALIGN', (0,0), (-1,-1), 'CENTER'),
#             ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            
#             # Zlúčenie buniek pre hlavičku
#             ('SPAN', (1,0), (2,0)),   # Dopoludnia
#             ('SPAN', (3,0), (4,0)),   # Odpoludnia
#             ('SPAN', (6,0), (10,0)),  # Príplatky
#             ('SPAN', (11,0), (12,0)), # Neodpr. hod.
            
#             # Fonty pre hlavičku
#             ('FONTNAME', (0,0), (-1,1), self.custom_font_bold),
            
#             # Spolu riadok
#             ('FONTNAME', (0,-1), (-1,-1), self.custom_font_bold),
#             ('BACKGROUND', (0,-1), (-1,-1), colors.lightgrey),
#         ]))
        
#         self.elements.append(t)
#         self.elements.append(Spacer(1, 5*mm))

#         # Poznámky
#         if self.notes_list:
#              notes_text = "Poznámky: " + "; ".join(self.notes_list)
#              self.elements.append(Paragraph(notes_text, self.styles['CustomNormal']))
#              self.elements.append(Spacer(1, 5*mm))

#         # Pätička
#         transferred = float(self.employee.initial_hours_balance or 0.0) if self.employee else 0.0
#         work_days_count = sum(1 for d in range(1, days_in_month + 1) if date(self.year, self.month, d).weekday() < 5)
#         fund = work_days_count * 7.5
        
#         balance = total_worked - fund
#         final = transferred + balance

#         footer_data = [
#             ["Prenesené hodiny", "FPČ", "Odpracované hod.\ncelkom +D+PN+P", "+ / - za mesiac", "Konečný stav\nhod. + -"],
#             [f"{transferred:+.1f}", f"{fund}", f"{total_worked:.1f}", f"{balance:+.1f}", f"{final:+.1f}"]
#         ]
        
#         footer_table = Table(footer_data, colWidths=[30*mm]*5)
#         footer_table.setStyle(TableStyle([
#             ('GRID', (0,0), (-1,-1), 0.5, colors.black),
#             ('ALIGN', (0,0), (-1,-1), 'CENTER'),
#             ('FONTNAME', (0,0), (-1,-1), self.custom_font_bold),
#         ]))
#         self.elements.append(footer_table)
#         self.elements.append(Spacer(1, 15*mm))
        
#         # Podpisy
#         signatures = [["Vyhotovil: ........................", "Kontroloval: ........................", "Schválil: ........................"]]
#         sig_table = Table(signatures, colWidths=[90*mm, 90*mm, 90*mm])
#         sig_table.setStyle(TableStyle([
#             ('ALIGN', (0,0), (-1,-1), 'LEFT'),
#             ('FONTNAME', (0,0), (-1,-1), self.custom_font),
#         ]))
#         self.elements.append(sig_table)

#         self.doc.build(self.elements)
        
#         pdf = self.buffer.getvalue()
#         self.buffer.close()
        
#         response = HttpResponse(content_type='application/pdf')
#         filename = f"Vykaz_{self.employee.last_name}_{self.month}_{self.year}.pdf" if self.employee else "Vykaz.pdf"
#         response['Content-Disposition'] = f'attachment; filename="{filename}"'
#         response.write(pdf)
#         return response


class AttendancePdfExporter:
    def __init__(self, employee_id, year, month):
        self.employee_id = employee_id
        self.year = year
        self.month = month
        
        # Ošetrenie, ak zamestnanec neexistuje
        try:
            self.employee = Employees.objects.get(pk=employee_id)
        except Employees.DoesNotExist:
             self.employee = None

        # --- REGISTRÁCIA FONTU (Arial) ---
        self.custom_font = 'Helvetica'
        self.custom_font_bold = 'Helvetica-Bold'
        
        try:
            fonts_dir = os.path.join(os.environ.get('WINDIR', 'C:\\Windows'), 'Fonts')
            arial_path = os.path.join(fonts_dir, 'arial.ttf')
            arialbd_path = os.path.join(fonts_dir, 'arialbd.ttf')
            
            if os.path.exists(arial_path) and os.path.exists(arialbd_path):
                pdfmetrics.registerFont(TTFont('ArialCustom', arial_path))
                pdfmetrics.registerFont(TTFont('ArialBoldCustom', arialbd_path))
                self.custom_font = 'ArialCustom'
                self.custom_font_bold = 'ArialBoldCustom'
        except Exception:
            pass
        
        # Nastavenie štýlov
        self.styles = getSampleStyleSheet()
        self.styles.add(ParagraphStyle(name='CustomNormal', parent=self.styles['Normal'], fontName=self.custom_font, fontSize=10))
        self.styles.add(ParagraphStyle(name='CustomTitle', parent=self.styles['Title'], fontName=self.custom_font_bold, alignment=1)) # 1=Center
        
        self.buffer = BytesIO()
        
        # --- ZMENA NA PORTRAIT (VÝŠKA) ---
        self.doc = SimpleDocTemplate(
            self.buffer, 
            pagesize=A4,  # <-- Tu je zmena (bez landscape)
            rightMargin=10*mm, leftMargin=10*mm, 
            topMargin=10*mm, bottomMargin=10*mm
        )
        self.elements = []
        self.notes_list = []

    def _format_time(self, t):
        if not t: return ""
        return t.strftime("%H,%M")

    def _format_duration(self, seconds):
        if not seconds: return ""
        hours = seconds / 3600.0
        if hours % 1 == 0:
            return f"{int(hours)}"
        return f"{hours:.1f}".replace(".", ",")

    def generate_response(self):
        # 1. Zber Dát
        days_in_month = monthrange(self.year, self.month)[1]
        start_date = date(self.year, self.month, 1)
        end_date = date(self.year, self.month, days_in_month)
        
        attendances = Attendance.objects.filter(
            user_id=self.employee_id,
            date__range=(start_date, end_date)
        ).select_related('type_shift').order_by('date')
        
        att_by_date = {a.date: a for a in attendances}
        holidays = set(val[0] for val in CalendarDay.get_slovak_holidays(self.year))

        slovak_months = ["", "Január", "Február", "Marec", "Apríl", "Máj", "Jún", 
                         "Júl", "August", "September", "Október", "November", "December"]

        # 2. Hlavička dokumentu
        header_text = f"<b>Centrum pre deti a rodiny Poprad</b><br/>Výkaz pedagogických pracovníkov, pomocných vychovávateľov a ostatných pracovníkov za"
        self.elements.append(Paragraph(header_text, self.styles['CustomTitle']))
        self.elements.append(Spacer(1, 5*mm))

        # Úprava šírky tabuľky s info (aby sa zmestila na A4 Portrait)
        # Celková šírka cca 190mm
        info_data = [
            [f"Mesiac / Rok: {slovak_months[self.month]} {self.year}", 
             f"Meno: {self.employee.last_name} {self.employee.first_name}" if self.employee else "", 
             f"Os.č.: {self.employee.personal_number or ''}" if self.employee else ""]
        ]
        info_table = Table(info_data, colWidths=[50*mm, 90*mm, 50*mm])
        info_table.setStyle(TableStyle([
            ('FONTNAME', (0,0), (-1,-1), self.custom_font_bold), 
            ('FONTSIZE', (0,0), (-1,-1), 10),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ]))
        self.elements.append(info_table)
        self.elements.append(Spacer(1, 5*mm))

        # 3. TABUĽKA - HLAVIČKA
        table_data = []
        
        headers_row1 = ["", "Dopoludnia", "", "Odpoludnia", "", "", "Príplatky", "", "", "", "", "Neodpr. hod.", ""]
        headers_row2 = ["Deň", "Príchod", "Odchod", "Príchod", "Odchod", "Odp.hod.", "Sob.", "Ned.", "Iná\nčinn.", "Sviatok", "Nočná", "Dovol.", "PN"]
        
        table_data.append(headers_row1)
        table_data.append(headers_row2)

        total_worked = 0.0
        sum_sob, sum_ned, sum_ina, sum_sviatok, sum_noc, sum_dov, sum_pn = 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0
        
        # 4. Riadky dní
        for day in range(1, days_in_month + 1):
            curr_date = date(self.year, self.month, day)
            att = att_by_date.get(curr_date)
            row = [f"{day}."] + [""] * 12
            
            if att:
                s_time = att.custom_start
                e_time = att.custom_end
                duration = 0.0
                
                if att.type_shift and att.type_shift.duration_time:
                     duration = float(att.type_shift.duration_time)
                elif s_time and e_time:
                    dt_s = datetime.combine(curr_date, s_time)
                    dt_e = datetime.combine(curr_date, e_time)
                    if dt_e < dt_s: dt_e += timedelta(days=1)
                    if e_time == time(0,0) and s_time != e_time: dt_e += timedelta(days=1)
                    duration = (dt_e - dt_s).total_seconds() / 3600.0

                ABSENCE_IDS = [21, 23] 

                if att.type_shift and att.type_shift.id not in ABSENCE_IDS:
                    if s_time and e_time:
                        NOON = time(12, 0)
                        if s_time < NOON:
                            row[1] = self._format_time(s_time)
                        else:
                            row[3] = self._format_time(s_time)

                        is_overnight = e_time < s_time
                        if is_overnight or e_time >= NOON:
                            row[4] = self._format_time(e_time)
                        else:
                            row[2] = self._format_time(e_time)

                    if duration > 0:
                        row[5] = self._format_duration(duration * 3600)
                        total_worked += duration

                if curr_date.weekday() == 5: 
                    row[6] = self._format_duration(duration * 3600)
                    sum_sob += duration
                elif curr_date.weekday() == 6: 
                    row[7] = self._format_duration(duration * 3600)
                    sum_ned += duration
                
                if att.type_shift and att.type_shift.id == 22:
                    row[8] = self._format_duration(duration * 3600)
                    sum_ina += duration
                
                if curr_date in holidays and att.type_shift.id not in ABSENCE_IDS:
                     row[9] = self._format_duration(duration * 3600)
                     sum_sviatok += duration

                if att.type_shift and att.type_shift.id == 20:
                     row[10] = self._format_duration(duration * 3600)
                     sum_noc += duration
                
                if att.type_shift and att.type_shift.id == 21:
                    row[11] = "D"
                    sum_dov += 7.5 
                
                if att.type_shift and att.type_shift.id == 23:
                    row[12] = "PN"
                    sum_pn += 7.5
                
                if att.note:
                    self.notes_list.append(f"{day}. {self.month}. - {att.note}")

            table_data.append(row)

        row_total = ["Spolu", "", "", "", "", 
                     self._format_duration(total_worked * 3600),
                     self._format_duration(sum_sob * 3600),
                     self._format_duration(sum_ned * 3600),
                     self._format_duration(sum_ina * 3600),
                     self._format_duration(sum_sviatok * 3600),
                     self._format_duration(sum_noc * 3600),
                     self._format_duration(sum_dov * 3600) if sum_dov else "",
                     self._format_duration(sum_pn * 3600) if sum_pn else ""]
        table_data.append(row_total)

        # --- NOVÉ ŠÍRKY STĹPCOV PRE PORTRAIT (VÝŠKA) ---
        # Celková šírka A4 = 210mm, okraje = 20mm, zostáva 190mm
        # Rozdelenie: 
        # Deň: 10mm
        # Časy (4 stĺpce): 16mm * 4 = 64mm
        # Zvyšné (8 stĺpcov): 13mm * 8 = 104mm
        # Spolu: 10 + 64 + 104 = 178mm (zmestí sa do 190mm)
        col_widths = [10*mm] + [16*mm]*4 + [13*mm]*8
        
        t = Table(table_data, colWidths=col_widths, repeatRows=2)
        
        t.setStyle(TableStyle([
            ('GRID', (0,0), (-1,-1), 0.5, colors.black),
            ('FONTNAME', (0,0), (-1,-1), self.custom_font),
            ('FONTSIZE', (0,0), (-1,-1), 8),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            
            ('SPAN', (1,0), (2,0)),
            ('SPAN', (3,0), (4,0)),
            ('SPAN', (6,0), (10,0)),
            ('SPAN', (11,0), (12,0)),
            
            ('FONTNAME', (0,0), (-1,1), self.custom_font_bold),
            ('FONTNAME', (0,-1), (-1,-1), self.custom_font_bold),
            ('BACKGROUND', (0,-1), (-1,-1), colors.lightgrey),
        ]))
        
        self.elements.append(t)
        self.elements.append(Spacer(1, 5*mm))

        if self.notes_list:
             notes_text = "Poznámky: " + "; ".join(self.notes_list)
             self.elements.append(Paragraph(notes_text, self.styles['CustomNormal']))
             self.elements.append(Spacer(1, 5*mm))

        transferred = float(self.employee.initial_hours_balance or 0.0) if self.employee else 0.0
        work_days_count = sum(1 for d in range(1, days_in_month + 1) if date(self.year, self.month, d).weekday() < 5)
        fund = work_days_count * 7.5
        balance = total_worked - fund
        final = transferred + balance

        footer_data = [
            ["Prenesené hodiny", "FPČ", "Odpracované hod.\ncelkom +D+PN+P", "+ / - za mesiac", "Konečný stav\nhod. + -"],
            [f"{transferred:+.1f}", f"{fund}", f"{total_worked:.1f}", f"{balance:+.1f}", f"{final:+.1f}"]
        ]
        
        # Úprava šírky pätičky pre Portrait
        footer_table = Table(footer_data, colWidths=[38*mm]*5)
        footer_table.setStyle(TableStyle([
            ('GRID', (0,0), (-1,-1), 0.5, colors.black),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME', (0,0), (-1,-1), self.custom_font_bold),
            ('FONTSIZE', (0,0), (-1,-1), 8),
        ]))
        self.elements.append(footer_table)
        self.elements.append(Spacer(1, 15*mm))
        
        # Úprava šírky podpisov pre Portrait
        signatures = [["Vyhotovil: ........................", "Kontroloval: ........................", "Schválil: ........................"]]
        sig_table = Table(signatures, colWidths=[63*mm, 63*mm, 63*mm])
        sig_table.setStyle(TableStyle([
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('FONTNAME', (0,0), (-1,-1), self.custom_font),
            ('FONTSIZE', (0,0), (-1,-1), 9),
        ]))
        self.elements.append(sig_table)

        self.doc.build(self.elements)
        
        pdf = self.buffer.getvalue()
        self.buffer.close()
        
        response = HttpResponse(content_type='application/pdf')
        filename = f"Vykaz_{self.employee.last_name}_{self.month}_{self.year}.pdf" if self.employee else "Vykaz.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response.write(pdf)
        return response


from reportlab.lib import colors
from reportlab.lib.pagesizes import A5, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from io import BytesIO
from datetime import date, timedelta
from calendar import monthrange
import os
from django.http import HttpResponse

# Importy modelov
from .models import PlannedShifts, Employees, CalendarDay

class VacationFormExporter:
    def __init__(self, employee_id, year, month):
        self.employee_id = employee_id
        self.year = year
        self.month = month
        
        # Načítanie zamestnanca
        try:
            self.employee = Employees.objects.get(pk=employee_id)
        except Employees.DoesNotExist:
             self.employee = None

        # --- REGISTRÁCIA FONTOV ---
        self.custom_font = 'Helvetica'
        self.custom_font_bold = 'Helvetica-Bold'
        
        try:
            fonts_dir = os.path.join(os.environ.get('WINDIR', 'C:\\Windows'), 'Fonts')
            arial_path = os.path.join(fonts_dir, 'arial.ttf')
            arialbd_path = os.path.join(fonts_dir, 'arialbd.ttf')
            
            if os.path.exists(arial_path) and os.path.exists(arialbd_path):
                pdfmetrics.registerFont(TTFont('ArialCustom', arial_path))
                pdfmetrics.registerFont(TTFont('ArialBoldCustom', arialbd_path))
                self.custom_font = 'ArialCustom'
                self.custom_font_bold = 'ArialBoldCustom'
        except Exception:
            pass
        
        # Nastavenie štýlov
        self.styles = getSampleStyleSheet()
        self.styles.add(ParagraphStyle(name='VacationTitle', parent=self.styles['Normal'], fontName=self.custom_font_bold, fontSize=14, alignment=TA_CENTER, spaceAfter=2*mm))
        self.styles.add(ParagraphStyle(name='VacationText', parent=self.styles['Normal'], fontName=self.custom_font, fontSize=10))
        self.styles.add(ParagraphStyle(name='VacationLabel', parent=self.styles['Normal'], fontName=self.custom_font, fontSize=7, textColor=colors.black))

        self.buffer = BytesIO()
        
        # A5 Landscape
        self.doc = SimpleDocTemplate(
            self.buffer, 
            pagesize=landscape(A5),
            rightMargin=10*mm, leftMargin=10*mm, 
            topMargin=10*mm, bottomMargin=10*mm
        )
        self.elements = []

    def _draw_border(self, canvas, doc):
        """Vykreslí rámček okolo celej strany"""
        canvas.saveState()
        canvas.setStrokeColor(colors.black)
        canvas.setLineWidth(2)
        width, height = landscape(A5)
        margin = 5*mm
        canvas.rect(margin, margin, width - 2*margin, height - 2*margin)
        canvas.restoreState()

    def _get_working_days_count(self, start_date, end_date):
        count = 0
        curr = start_date
        holidays = set(val[0] for val in CalendarDay.get_slovak_holidays(self.year))
        while curr <= end_date:
            if curr.weekday() < 5 and curr not in holidays:
                count += 1
            curr += timedelta(days=1)
        return count

    def _group_consecutive_dates(self, shifts):
        if not shifts: return []
        groups = []
        current_group = [shifts[0]]
        for i in range(1, len(shifts)):
            prev_date = shifts[i-1].date
            curr_date = shifts[i].date
            if curr_date == prev_date + timedelta(days=1):
                current_group.append(shifts[i])
            else:
                groups.append(current_group)
                current_group = [shifts[i]]
        if current_group: groups.append(current_group)
        return groups

    def _draw_form(self, group):
        start_date = group[0].date
        end_date = group[-1].date
        work_days = self._get_working_days_count(start_date, end_date)
        
        # --- 1. NADPIS ---
        self.elements.append(Paragraph("DOVOLENKA", self.styles['VacationTitle']))
        self.elements.append(Spacer(1, 2*mm))

        # --- 2. HLAVNÝ RÁMEC S ÚDAJMI ---
        emp_name = f"{self.employee.last_name} {self.employee.first_name}" if self.employee else ""
        emp_num = f"{self.employee.personal_number}" if self.employee else ""
        
        # Riadok 1: Meno a Osobné číslo
        data_top = [
            [Paragraph("Priezvisko, meno, titul", self.styles['VacationLabel']), Paragraph("Osobné číslo", self.styles['VacationLabel'])],
            [Paragraph(f"{emp_name}", self.styles['VacationText']), Paragraph(f"{emp_num}", self.styles['VacationText'])]
        ]
        t_top = Table(data_top, colWidths=[130*mm, 50*mm])
        t_top.setStyle(TableStyle([
            ('LINEBELOW', (0,1), (-1,1), 1, colors.black),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('LEFTPADDING', (0,0), (-1,-1), 0),
        ]))
        self.elements.append(t_top)
        
        # Riadok 2: Útvar
        data_dept = [
            [
                Paragraph("Útvar: Centrum pre deti a rodiny Poprad, Pavlovova 4375/", self.styles['VacationText']),
                Paragraph("Číslo útvaru ...........", self.styles['VacationText'])
            ]
        ]
        t_dept = Table(data_dept, colWidths=[130*mm, 50*mm])
        t_dept.setStyle(TableStyle([
            ('LINEBELOW', (0,0), (-1,-1), 1, colors.black),
            ('LEFTPADDING', (0,0), (-1,-1), 0),
            ('ALIGN', (1,0), (1,0), 'RIGHT'),
            ('VALIGN', (0,0), (-1,-1), 'BOTTOM'),
        ]))
        self.elements.append(t_dept)
        
        self.elements.append(Spacer(1, 5*mm))

        # --- 3. DÁTUMY ---
        self.elements.append(Paragraph(f"Žiada o dovolenku na zotavenie za kalendárny rok {self.year}", self.styles['VacationText']))
        self.elements.append(Spacer(1, 3*mm))
        
        s_str = start_date.strftime("%d. %m. %Y")
        e_str = end_date.strftime("%d. %m. %Y")
        
        data_dates = [
            [f"od  {s_str}", f"do  {e_str}", f"vrátane, t.j.  {work_days}  pracovných dní"]
        ]
        t_dates = Table(data_dates, colWidths=[50*mm, 50*mm, 80*mm])
        t_dates.setStyle(TableStyle([
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('FONTNAME', (0,0), (-1,-1), self.custom_font),
            ('FONTSIZE', (0,0), (-1,-1), 10),
            ('LEFTPADDING', (0,0), (-1,-1), 0),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ]))
        self.elements.append(t_dates)
        
        self.elements.append(Paragraph("Miesto pobytu na dovolenke: .....................................................................................", self.styles['VacationText']))
        self.elements.append(Spacer(1, 5*mm))

        # --- 4. PODPIS (UPRAVENÉ) ---
        today_str = date.today().strftime("%d. %m. %Y")
        data_sig = [
            [f"{today_str}", "........................................................"],
            ["Dátum", "podpis pracovníka"]
        ]
        t_sig = Table(data_sig, colWidths=[60*mm, 80*mm])
        t_sig.setStyle(TableStyle([
            # Zarovnanie na stred pre oba stĺpce (Dátum aj Podpis)
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            # Odstránené čiary (LINEABOVE)
            ('FONTSIZE', (0,1), (-1,1), 7), # Menší font pre popisky
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ]))
        self.elements.append(t_sig)
        self.elements.append(Spacer(1, 5*mm))

        # --- 5. SCHVAĽOVACIA TABUĽKA ---
        data_approval = [
            ["", "Dátum", "Ved. Útvaru", "Pers. odbor"],
            ["Schválil", "", "", ""],
            ["Skutočný nástup dovolenky", "", "", ""],
            ["Nástup do zam. po dovolenke", "", "", ""],
            [f"Z tejto dovolenky sa skutočne čerpalo: {work_days} prac. dní", "", "", ""]
        ]
        
        t_app = Table(data_approval, colWidths=[80*mm, 25*mm, 35*mm, 35*mm], rowHeights=8*mm)
        t_app.setStyle(TableStyle([
            ('GRID', (0,0), (-1,-1), 1, colors.black),
            ('FONTNAME', (0,0), (-1,-1), self.custom_font_bold),
            ('FONTSIZE', (0,0), (-1,-1), 8),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('ALIGN', (0,1), (0,-1), 'LEFT'),
            ('SPAN', (0,4), (1,4)),
        ]))
        self.elements.append(t_app)
        
        self.elements.append(Spacer(1, 2*mm))
        self.elements.append(Paragraph("Poznámky o mzdových nárokoch, neodovzdaných nástrojoch uveďte na zadnej strane.", self.styles['VacationLabel']))

        self.elements.append(PageBreak())

    def generate_response(self):
        VACATION_IDS = [21, 5] 

        days_in_month = monthrange(self.year, self.month)[1]
        start_date = date(self.year, self.month, 1)
        end_date = date(self.year, self.month, days_in_month)

        shifts = PlannedShifts.objects.filter(
            user_id=self.employee_id,
            date__range=(start_date, end_date),
            type_shift__id__in=VACATION_IDS
        ).order_by('date')

        grouped_shifts = self._group_consecutive_dates(list(shifts))

        if not grouped_shifts:
            self.elements.append(Paragraph("V tomto období nebola nájdená žiadna plánovaná dovolenka.", self.styles['VacationText']))
            self.doc.build(self.elements, onFirstPage=self._draw_border, onLaterPages=self._draw_border)
            pdf = self.buffer.getvalue()
            self.buffer.close()
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="Ziadna_dovolenka.pdf"'
            response.write(pdf)
            return response

        for group in grouped_shifts:
            self._draw_form(group)

        self.doc.build(self.elements, onFirstPage=self._draw_border, onLaterPages=self._draw_border)
        
        pdf = self.buffer.getvalue()
        self.buffer.close()
        
        response = HttpResponse(content_type='application/pdf')
        filename = f"Dovolenky_{self.employee.last_name}_{self.month}_{self.year}.pdf" if self.employee else "Dovolenky.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response.write(pdf)
        return response

class ExchangeFormExporter:
    def __init__(self, source_shift, target_shift):
        self.source_shift = source_shift
        self.target_shift = target_shift
        self.employee = source_shift.user
        
        # --- REGISTRÁCIA FONTOV (Rovnako ako v VacationFormExporter) ---
        self.custom_font = 'Helvetica'
        self.custom_font_bold = 'Helvetica-Bold'
        
        try:
            fonts_dir = os.path.join(os.environ.get('WINDIR', 'C:\\Windows'), 'Fonts')
            arial_path = os.path.join(fonts_dir, 'arial.ttf')
            arialbd_path = os.path.join(fonts_dir, 'arialbd.ttf')
            
            if os.path.exists(arial_path) and os.path.exists(arialbd_path):
                pdfmetrics.registerFont(TTFont('ArialCustom', arial_path))
                pdfmetrics.registerFont(TTFont('ArialBoldCustom', arialbd_path))
                self.custom_font = 'ArialCustom'
                self.custom_font_bold = 'ArialBoldCustom'
        except Exception:
            pass
        
        # Nastavenie štýlov
        self.styles = getSampleStyleSheet()
        # Nadpis (Väčší, Bold, Center)
        self.styles.add(ParagraphStyle(name='ExchangeTitle', parent=self.styles['Normal'], 
                                       fontName=self.custom_font_bold, fontSize=16, 
                                       alignment=TA_CENTER, spaceAfter=10*mm))
        # Bežný text
        self.styles.add(ParagraphStyle(name='ExchangeText', parent=self.styles['Normal'], 
                                       fontName=self.custom_font, fontSize=11, leading=14))
        # Bold text
        self.styles.add(ParagraphStyle(name='ExchangeLabel', parent=self.styles['Normal'], 
                                       fontName=self.custom_font_bold, fontSize=11, leading=14))

        self.buffer = BytesIO()
        
        # A4 Portrait (Na rozdiel od dovolenky, ktorá bola A5)
        self.doc = SimpleDocTemplate(
            self.buffer, 
            pagesize=A4,
            rightMargin=25*mm, leftMargin=25*mm, 
            topMargin=20*mm, bottomMargin=20*mm
        )
        self.elements = []

    def _get_duration_string(self):
        """Pomocná funkcia na formátovanie času"""
        if self.source_shift.custom_start and self.source_shift.custom_end:
            # Predpokladáme, že custom_start je time objekt alebo string
            s = str(self.source_shift.custom_start)[:5] # Oreže sekundy
            e = str(self.source_shift.custom_end)[:5]
            return f"{s} - {e}"
        return "Podľa rozpisu"

    def generate_response(self):
        # Dáta
        applicant_name = f"{self.employee.first_name} {self.employee.last_name}"
        substitute_name = f"{self.target_shift.user.first_name} {self.target_shift.user.last_name}"
        shift_date_str = self.source_shift.date.strftime("%d.%m.%Y")
        today_str = date.today().strftime("%d.%m.%Y")
        hours_str = self._get_duration_string()
        reason_text = self.source_shift.change_reason.name if self.source_shift.change_reason else "Výmena služby"

        # --- 1. HLAVIČKA (Meno a Dátum) ---
        # Tabuľka bez rámikov pre zarovnanie
        data_header = [
            [Paragraph(f"<b>Meno a priezvisko žiadateľa:</b> {applicant_name}", self.styles['ExchangeText'])],
            [Paragraph(f"<b>Dňa:</b> {today_str} <b>vo Veľkom Slavkove</b>", self.styles['ExchangeText'])]
        ]
        t_header = Table(data_header, colWidths=[160*mm])
        t_header.setStyle(TableStyle([
            ('LEFTPADDING', (0,0), (-1,-1), 0),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ]))
        self.elements.append(t_header)
        self.elements.append(Spacer(1, 15*mm))

        # --- 2. NADPIS ---
        self.elements.append(Paragraph("Žiadosť o povolenie výmeny služby", self.styles['ExchangeTitle']))
        self.elements.append(Spacer(1, 10*mm))

        # --- 3. HLAVNÝ TEXT ---
        # Používame f-string s HTML tagmi (<b> pre bold) v rámci Paragraphu
        text_body = f"""
        Dolepodpísaný/á žiadam vedenie CPDaR Poprad o povolenie výmeny služby dňa 
        <b>{shift_date_str}</b> v počte hodín <b>{hours_str}</b> 
        s pracovníkom/pracovníčkou: <b>{substitute_name}</b>.
        """
        self.elements.append(Paragraph(text_body, self.styles['ExchangeText']))
        self.elements.append(Spacer(1, 10*mm))

        # --- 4. DÔVOD ---
        self.elements.append(Paragraph("<b>Dôvod žiadosti:</b>", self.styles['ExchangeText']))
        self.elements.append(Spacer(1, 2*mm))
        self.elements.append(Paragraph(f"<i>{reason_text}</i>", self.styles['ExchangeText']))
        self.elements.append(Spacer(1, 15*mm))

        # --- 5. ODPRACOVANIE (Riadok s bodkami) ---
        self.elements.append(Paragraph(
            "Výkon služby v počte hodín odpracujem dňa ...............................................................", 
            self.styles['ExchangeText']
        ))
        self.elements.append(Spacer(1, 25*mm))

        # --- 6. PODPISY (Tabuľka) ---
        # 1. Riadok: Podpis žiadateľa (Vľavo)
        # 2. Riadok: Vedúci (Vľavo) a Riaditeľ (Vpravo)
        
        dots = "..............................................................."
        
        # Riadok 1: Žiadateľ
        data_sig_1 = [
            [dots, ""],
            ["Podpis žiadateľa", ""]
        ]
        t_sig_1 = Table(data_sig_1, colWidths=[80*mm, 80*mm])
        t_sig_1.setStyle(TableStyle([
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME', (0,1), (-1,1), self.custom_font),
            ('FONTSIZE', (0,1), (-1,1), 9),
            ('TOPPADDING', (0,1), (-1,1), 0),
        ]))
        self.elements.append(t_sig_1)
        self.elements.append(Spacer(1, 15*mm))

        # Riadok 2: Vedúci a Riaditeľ
        data_sig_2 = [
            [dots, dots],
            ["Podpis vedúceho výchovy", "Podpis riaditeľa CPDaR"]
        ]
        t_sig_2 = Table(data_sig_2, colWidths=[80*mm, 80*mm])
        t_sig_2.setStyle(TableStyle([
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME', (0,1), (-1,1), self.custom_font),
            ('FONTSIZE', (0,1), (-1,1), 9),
            ('TOPPADDING', (0,1), (-1,1), 0), # Aby bol text hneď pod bodkami
        ]))
        self.elements.append(t_sig_2)

        # --- GENERATE PDF ---
        self.doc.build(self.elements)
        
        pdf = self.buffer.getvalue()
        self.buffer.close()
        
        response = HttpResponse(content_type='application/pdf')
        filename = f"Ziadost_o_vymenu_{self.source_shift.id}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response.write(pdf)
        return response
#end