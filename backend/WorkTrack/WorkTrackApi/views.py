from rest_framework.exceptions import ValidationError
from .export import MonthlyRosterExporter, AttendancePdfExporter,VacationFormExporter
from rest_framework import viewsets, permissions, status, generics
from .models import Employees, TypeShift, Attendance, PlannedShifts, ChangeReason, CalendarDay
from .serializers import BulkRosterSerializer, EmployeesSerializer, TypeShiftSerializer, AttendanceSerializer, PlannedShiftsSerializer, ChangeReasonSerializers, CalendarDaySerializers
from .services import calculate_working_fund, calculate_worked_hours, calculate_saturday_sunday_hours, calculate_weekend_hours, calculate_holiday_hours, compare_worked_time_working_fund, calculate_total_hours_with_transfer, calculate_night_shift_hours, copy_monthly_plan, get_planned_monthly_summary
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.exceptions import PermissionDenied
from datetime import time, timedelta
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.views import APIView
from django.utils.dateparse import parse_time
from rest_framework.authtoken.views import ObtainAuthToken
from rest_framework.authtoken.models import Token
from rest_framework.authentication import TokenAuthentication
from .permissions import IsManagerOrReadOnly

# Import utility funkcii
from WorkTrackApi.utils.attendance_utils import (
    split_night_planned_shift,
    handle_night_shift,
    handle_start_shift_time,
    handle_end_shift_time
)

# ==========================================
# API VIEWS (Reports & Stats)
# ==========================================

class BaseWorkedHoursAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def calculate_hours(self, employee_id, year, month):
        raise NotImplementedError("Potomok musí implementovať calculate_hours")

    def get(self, request, employee_id, year, month):
        user = request.user
        if user.role == "worker" and user.id != employee_id:
            return Response({"detail": "Nemáš oprávnenie vidieť údaje iných."}, status=403)

        hours = self.calculate_hours(employee_id, year, month)
        return Response({"worked_hours": hours})

from .models import CalendarDay  # 👈 Nezabudnite importovať model!
from .services import calculate_working_fund,get_balances_up_to
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import permissions

class MonthlyBalancesAPIView(APIView):
    def get(self, request, year, month):
        # Zavoláme servisnú funkciu
        balances = get_balances_up_to(year, month)
        return Response(balances)

class WorkingFundAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, year: int, month: int):
        # 1. Existujúci výpočet fondu
        fund = calculate_working_fund(year, month)

        # 2. 👇 NOVÉ: Získanie zoznamu sviatkov pre daný mesiac
        holidays_query = CalendarDay.objects.filter(
            date__year=year, 
            date__month=month, 
            is_holiday=True  # Hľadáme len dni označené ako sviatok
        )
        
        # Vytvoríme jednoduchý zoznam čísel dní (napr. [1, 6, 15] pre január)
        # Frontend očakáva pole čísiel, aby vedel porovnať s day.dayNum
        holiday_days = [h.date.day for h in holidays_query]

        # 3. Vrátenie odpovede aj so zoznamom sviatkov
        return Response({
            "year": year,
            "month": month,
            "working_fund_hours": fund,
            "holidays": holiday_days  # 👈 Toto pole využije Angular na farbenie
        })
class WorkedHoursAPIView(BaseWorkedHoursAPIView):
    def calculate_hours(self, employee_id, year, month):
        return calculate_worked_hours(employee_id, year, month)

class SaturdaySundayHoursApiView(BaseWorkedHoursAPIView):
    def calculate_hours(self, employee_id, year, month):
        return calculate_saturday_sunday_hours(employee_id, year, month)

class WeekendHoursApiView(BaseWorkedHoursAPIView):
    def calculate_hours(self, employee_id, year, month):
        return calculate_weekend_hours(employee_id, year, month)

class CompareHoursApiView(BaseWorkedHoursAPIView):
    def calculate_hours(self, employee_id, year, month):
        return compare_worked_time_working_fund(employee_id, year, month)

class HolidayHoursApiView(BaseWorkedHoursAPIView):
    def calculate_hours(self, employee_id, year, month):
        return calculate_holiday_hours(employee_id, year, month)

class TotalHoursApiView(BaseWorkedHoursAPIView):
    def calculate_hours(self, employee_id, year, month):
        return calculate_total_hours_with_transfer(employee_id, year, month)

class NightShiftHoursApiView(BaseWorkedHoursAPIView):
    def calculate_hours(self, employee_id, year, month):
        return calculate_night_shift_hours(employee_id, year, month)

class DashboardView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        return Response({"message": f"Ahoj {request.user.username}, toto je tvoj dashboard"})

class TestView(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        return Response({"message": "OK"})

class ActiveUserListView(generics.ListAPIView):
    serializer_class = EmployeesSerializer
    def get_queryset(self):
         return Employees.objects.all().prefetch_related('planned_shifts')

class CurrentUserView(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        user = request.user
        serializer = EmployeesSerializer(user)
        return Response(serializer.data)

class CustomAuthToken(ObtainAuthToken):
    def post(self, request, *args, **kwargs):
        serializer = self.serializer_class(data=request.data, context={'request': request})
        serializer.is_valid(raise_exception=True)
        user = serializer.validated_data['user']
        token, _ = Token.objects.get_or_create(user=user)
        return Response({
            'token': token.key,
            'user_id': user.id,
            'username': user.username,
            'role': user.role
        })

# ==========================================
# VIEWSETS (CRUD Operácie)
# ==========================================

class EmployeesViewSet(viewsets.ModelViewSet):
    queryset = Employees.objects.all()
    serializer_class = EmployeesSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        user=self.request.user
        if user.role == 'worker':
            return Employees.objects.filter(id=user.id)
        return Employees.objects.all()
    
    def get_permissions(self):
        user = self.request.user
        if user.is_authenticated and user.role == 'worker':
            self.http_method_names = ['get', 'head', 'options']
        else:
            self.http_method_names = ['get', 'post', 'put', 'patch', 'delete', 'head', 'options']
        return super().get_permissions()

class TypeShiftViewSet(viewsets.ModelViewSet):
    queryset = TypeShift.objects.all()
    serializer_class = TypeShiftSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_permissions(self):
        user = self.request.user
        if user.role == 'worker':
            self.http_method_names = ['get', 'head', 'options']
        else:
            self.http_method_names = ['get', 'post', 'put', 'patch', 'delete', 'head', 'options']
        return super().get_permissions()

    def perform_create(self, serializer):
        user = self.request.user
        if user.role == 'worker':
            raise PermissionDenied("Nemáte oprávnenie vytvárať smeny.")
        serializer.save()

class AttendanceViewSet(viewsets.ModelViewSet):
    queryset = Attendance.objects.all()
    serializer_class = AttendanceSerializer
    permission_classes = [permissions.IsAuthenticated]
    @action(detail=False, methods=['get'], url_path='export-attendance-pdf')
    def export_attendance_pdf(self, request):
        """
        Generuje PDF Výkaz pedagogického pracovníka na základe dochádzky.
        """
        try:
            year = int(request.query_params.get('year'))
            month = int(request.query_params.get('month'))
            user_id = int(request.query_params.get('user_id'))
        except (TypeError, ValueError):
            return Response({"error": "Chýbajú parametre year, month alebo user_id"}, status=400)
            
        exporter = AttendancePdfExporter(user_id, year, month)
        return exporter.generate_response()
    def get_queryset(self):
        user = self.request.user
        if user.role == 'worker':
            return Attendance.objects.filter(user=user)
        return Attendance.objects.all()

    def perform_create(self, serializer):
        # 1. Uložíme základný záznam
        instance = serializer.save()

        # 2. Logika naviazaná na typ smeny
        if instance.type_shift:
            # A) Nočná smena (ID 20) -> vytvorenie 2. dňa
            if instance.type_shift.id == 20:
                handle_night_shift(instance)
            
            # B) Riešenie časov (extra smeny pri skoršom príchode/neskoršom odchode)
            handle_start_shift_time(instance)
            handle_end_shift_time(instance)

    def perform_update(self, serializer):
        # Pri update tiež kontrolujeme časy
        if 'user' not in serializer.validated_data:
            instance = serializer.save(user=self.get_object().user)
        else:
            instance = serializer.save()

        if instance.type_shift:
            # Pri zmene času sa môže aktivovať extra smena
            handle_start_shift_time(instance)
            handle_end_shift_time(instance)

    def perform_destroy(self, instance):
        user = self.request.user
        if user.role == 'worker' and instance.user != user:
            raise PermissionDenied("Nemôžete zmazať cudzí záznam.")
        instance.delete()

# ==========================================
# PLANNER SHIFTS VIEWSET (Opravený)
# ==========================================
from django.db import transaction
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser, FormParser
from datetime import date, datetime, time
from calendar import monthrange
import csv
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ... tvoje importy ...


class PlannerShiftsViewSet(viewsets.ModelViewSet):
    queryset = PlannedShifts.objects.all()
    serializer_class = PlannedShiftsSerializer
    permission_classes = [IsManagerOrReadOnly]



    @action(detail=False, methods=['get'], url_path='export-vacation-forms')
    def export_vacation_forms(self, request):
        """
        Generuje PDF lístky na dovolenku na základe plánu.
        Prerušenie dátumov vytvorí nový lístok.
        """
        try:
            year = int(request.query_params.get('year'))
            month = int(request.query_params.get('month'))
            user_id = int(request.query_params.get('user_id'))
        except (TypeError, ValueError):
             # Ak nie sú parametre, skúsime default alebo error
             return Response({"error": "Chýbajú parametre year, month alebo user_id"}, status=400)
            
        exporter = VacationFormExporter(user_id, year, month)
        return exporter.generate_response()
    
    @action(detail=False, methods=['get'], url_path='export-complex-roster')
    def export_complex_roster(self, request):
        
        # 1. Získanie parametrov z URL (?year=2025&month=4)
        try:
            year = int(request.query_params.get('year', 2025))
            month = int(request.query_params.get('month', 4))
        except ValueError:
            return Response({"error": "Neplatný formát roku alebo mesiaca."}, status=400)
        
        # 2. Vytvorenie inštancie exportera
        # (Tu sa zinicializuje Workbook, načítajú štýly a konfigurácie)
        exporter = MonthlyRosterExporter(year, month)
        
        # 3. Spustenie generovania a vrátenie odpovede
        # (Metóda generate_response() spraví všetku prácu:
        #  - načíta dáta z DB
        #  - skontroluje či existujú (ak nie, vyhodí 400)
        #  - vykreslí hlavičku, tabuľku, sumáre a legendu
        #  - vráti hotový HttpResponse s Excel súborom)
        return exporter.generate_response()
  
    @action(detail=False, methods=['post'], url_path='save-roster-matrix')
    def save_roster_matrix(self, request):
        """
        Tento endpoint slúži ako 'Backend Formulár' pre uloženie celého rozdeľovníka naraz.
        Frontend pošle JSON: { "year": 2025, "month": 4, "shifts": [...] }
        """
        serializer = BulkRosterSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        data = serializer.validated_data
        year = data['year']
        month = data['month']
        shifts_data = data['shifts']
        
        # Získame rozsah mesiaca pre bezpečné mazanie/úpravu
        start_date = date(year, month, 1)
        _, last_day = monthrange(year, month)
        end_date = date(year, month, last_day)

        saved_count = 0
        
        try:
            with transaction.atomic():
                # Voliteľné: Vymazať existujúce plány pre dotknutých užívateľov v tomto mesiaci?
                # Alebo použijeme update_or_create pre "inteligentné" prepísanie.
                # Tu používam prístup update_or_create pre každú bunku.
                
                for item in shifts_data:
                    user_id = item['user_id']
                    shift_date = item['date']
                    
                    # Validácia, či dátum patrí do mesiaca (bezpečnosť)
                    if shift_date.year != year or shift_date.month != month:
                        continue

                    # Ak type_shift_id je None, znamená to zmazanie smeny (vyčistenie bunky)
                    if item.get('type_shift_id') is None:
                        PlannedShifts.objects.filter(
                            user_id=user_id, 
                            date=shift_date
                        ).delete()
                        continue

                    # Príprava dát na uloženie
                    defaults = {
                        'type_shift_id': item['type_shift_id'],
                        'custom_start': item.get('custom_start'),
                        'custom_end': item.get('custom_end'),
                        'note': item.get('note', ''),
                        'hidden': False
                    }
                    
                    # Logika pre automatické časy, ak nie sú zadané (podľa TypeShift)
                    if not defaults['custom_start']:
                         ts = TypeShift.objects.get(id=item['type_shift_id'])
                         defaults['custom_start'] = ts.start_time
                         defaults['custom_end'] = ts.end_time

                    # Update alebo Create
                    shift, created = PlannedShifts.objects.update_or_create(
                        user_id=user_id,
                        date=shift_date,
                        defaults=defaults
                    )
                    
                    # Aplikovanie logiky pre nočné smeny (z tvojho pôvodného kódu)
                    if shift.type_shift and shift.type_shift.id == 20:
                        split_night_planned_shift(shift)
                    
                    saved_count += 1

            return Response({"detail": f"Úspešne uložených/aktualizovaných {saved_count} záznamov."}, status=200)
        
        except Exception as e:
            return Response({"error": str(e)}, status=500)

    # @action(detail=False, methods=['post'], parser_classes=[MultiPartParser, FormParser], url_path='import-csv')
    # def import_roster_csv(self, request):
    #     """
    #     Parsuje tvoj špecifický Excel/CSV formát:
    #     Row 3: Mená zamestnancov (každých X stĺpcov)
    #     Col A: Dátum
    #     Cells: Hodnoty (napr. 12, 7, D, N)
    #     """
    #     file_obj = request.FILES.get('file')
    #     if not file_obj:
    #         return Response({"error": "Súbor chýba"}, status=400)

    #     try:
    #         # Načítanie CSV
    #         decoded_file = file_obj.read().decode('utf-8').splitlines()
    #         reader = csv.reader(decoded_file, delimiter=',') # alebo ';' podľa formátu
    #         rows = list(reader)
            
    #         # --- LOGIKA PARSOVANIA TVOJHO SÚBORU ---
    #         # 1. Nájdenie hlavičky s menami (Riadok 3 v tvojom CSV - index 2)
    #         header_row_idx = 2 
    #         header = rows[header_row_idx]
            
    #         # Mapovanie mien na user_id (Toto musíš prispôsobiť tvojej DB)
    #         # Príklad: "p.Polyak" -> user_id: 10
    #         employee_map = {} 
    #         # Tu prechádzame stĺpce a hľadáme mená. V tvojom CSV sú mená občasne.
    #         # Musíš si vytvoriť logiku, ktorý stĺpec patrí komu.
    #         # Zjednodušený príklad:
    #         current_user = None
    #         col_user_map = {} # {column_index: user_name}
            
    #         for i, cell in enumerate(header):
    #             if cell and cell.strip():
    #                 current_user = cell.strip() # Napr "p.Polyak"
    #                 # Skúsime nájsť Usera v DB podľa mena
    #                 # user = Employees.objects.filter(last_name__icontains=current_user.split('.')[1]).first()
    #                 # if user: col_user_map[i] = user.id
            
    #         # 2. Iterácia cez dni (od riadku 4 - index 3)
    #         created_count = 0
    #         with transaction.atomic():
    #             for i in range(3, len(rows)):
    #                 row = rows[i]
    #                 if not row or not row[0]: continue # Preskočiť prázdne riadky
                    
    #                 day_str = row[0].replace('.', '') # "1." -> "1"
    #                 try:
    #                     day = int(day_str)
    #                 except ValueError:
    #                     continue # Nie je to riadok s dňom
                        
    #                 # Dátum (Musíme vedieť rok a mesiac - buď z inputu alebo zo súboru)
    #                 # Predpokladám, že rok/mesiac pošleš v query params alebo sú v názve súboru
    #                 target_year = int(request.query_params.get('year', 2025))
    #                 target_month = int(request.query_params.get('month', 4))
    #                 current_date = date(target_year, target_month, day)

    #                 # Teraz prejdeme stĺpce priradené userom
    #                 # V tvojom CSV sa zdá, že pre každého usera je viac stĺpcov (SN, Do, Pp...)
    #                 # Musíš presne definovať, ktorý stĺpec obsahuje typ smeny.
    #                 # Podľa snippetu: Polyak má stĺpce 2-7, hodnota je v stĺpci 2?
                    
    #                 # *PSEUDOKÓD PRE SPRACOVANIE BUNKY*:
    #                 # value = row[user_column_index]
    #                 # shift_type = map_csv_value_to_shift_type(value)
    #                 # if shift_type:
    #                 #    PlannedShifts.objects.update_or_create(...)
            
    #         return Response({"detail": "Import dokončený (Logiku parsovania treba doladiť podľa presných stĺpcov)."}, status=200)

    #     except Exception as e:
    #         return Response({"error": f"Chyba pri spracovaní: {str(e)}"}, status=500)





    def get_queryset(self):
        user = self.request.user
        if user.is_anonymous:
            return PlannedShifts.objects.none()

        if user.role == 'worker':
            return PlannedShifts.objects.filter(user=user, hidden=False)
        
        # Admin/Manager filtre
        queryset = PlannedShifts.objects.filter(hidden=False)
        month = self.request.query_params.get('month')
        year = self.request.query_params.get('year')
        user_id = self.request.query_params.get('user')
        
        if month and year:
            queryset = queryset.filter(date__year=year, date__month=month)
        if user_id:
            queryset = queryset.filter(user_id=user_id)
            
        return queryset

    def create(self, request, *args, **kwargs):
        # 1. Podpora pre Bulk Create (zoznam objektov)
        is_many = isinstance(request.data, list)
        
        # 2. Serializácia
        serializer = self.get_serializer(data=request.data, many=is_many)
        serializer.is_valid(raise_exception=True)
        
        # 3. Uloženie (volá perform_create)
        self.perform_create(serializer)
        
        # 4. VRÁTENIE ODPOVEDE (Toto opravuje chybu 500)
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    def perform_create(self, serializer):
        # Uloženie dát
        created_data = serializer.save()

        # Logika pre rozdelenie nočnej smeny (ID 20)
        # Musíme zistiť, či sme vytvorili jeden objekt alebo zoznam
        if isinstance(created_data, list):
            for shift in created_data:
                if shift.type_shift and shift.type_shift.id == 20:
                    split_night_planned_shift(shift)
        else:
            if created_data.type_shift and created_data.type_shift.id == 20:
                split_night_planned_shift(created_data)

    def perform_update(self, serializer):
        planned_shift = serializer.save()
        # Aj pri úprave kontrolujeme nočnú smenu
        if planned_shift.type_shift and planned_shift.type_shift.id == 20:
            split_night_planned_shift(planned_shift)

    # --- CUSTOM ACTIONS ---

    @action(detail=False, methods=['post'])
    def copy_plan(self, request):
        s_year = request.data.get('source_year')
        s_month = request.data.get('source_month')
        t_year = request.data.get('target_year')
        t_month = request.data.get('target_month')
        user_id = request.data.get('user_id')

        if not all([s_year, s_month, t_year, t_month]):
            return Response({"error": "Chýbajú povinné údaje."}, status=400)

        try:
            result = copy_monthly_plan(int(s_year), int(s_month), int(t_year), int(t_month), user_id)
            return Response({"detail": "Plán skopírovaný.", "stats": result}, status=200)
        except Exception as e:
            return Response({"error": str(e)}, status=500)

    @action(detail=False, methods=['get'], url_path='summary/(?P<user_id>\d+)/(?P<year>\d+)/(?P<month>\d+)')
    def monthly_summary(self, request, user_id=None, year=None, month=None):
        if not user_id or not year or not month:
            return Response({"error": "Chýbajú parametre"}, status=400)
        try:
            summary_data = get_planned_monthly_summary(int(user_id), int(year), int(month))
            return Response(summary_data)
        except Exception as e:
            return Response({"error": str(e)}, status=500)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.hidden = True
        instance.save()
        return Response({"detail": "Smena bola skrytá."}, status=status.HTTP_200_OK)

class ChangeReasonViewSet(viewsets.ModelViewSet):
    queryset = ChangeReason.objects.all()
    serializer_class = ChangeReasonSerializers
    permission_classes = [permissions.IsAuthenticated]

    def get_permissions(self):
        user = self.request.user
        if user.role == 'worker':
            self.http_method_names = ['get', 'head', 'options']
        else:
            self.http_method_names = ['get', 'post', 'put', 'patch', 'delete', 'head', 'options']
        return super().get_permissions()

    def perform_create(self, serializer):
        user = self.request.user
        if user.role == 'worker':
            raise PermissionDenied("Nemáte oprávnenie vytvárať tieto dôvody.")
        serializer.save()

class CalendarDayViewSet(viewsets.ModelViewSet):
    queryset = CalendarDay.objects.all()
    serializer_class = CalendarDaySerializers
    permission_classes = [permissions.IsAuthenticated]

    def get_permissions(self):
        user = self.request.user
        # Tu bola v pôvodnom kóde malá logická chyba 'and', opravené na štandard
        if user.role == 'worker': 
            self.http_method_names = ['get', 'head', 'options']
        else:
            self.http_method_names = ['get', 'post', 'put', 'patch', 'delete', 'head', 'options']
        return super().get_permissions()

    def perform_create(self, serializer):
        user = self.request.user
        if user.role == 'worker':
            raise PermissionDenied("Nemáte oprávnenie vytvárať tieto dôvody.")
        serializer.save()

class PlannedHoursSummaryView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, employee_id, year, month):
        # Zavoláme funkciu zo services.py
        data = get_planned_monthly_summary(employee_id, year, month)
        
        # Vrátime výsledok ako JSON
        return Response(data)
    


def map_csv_value_to_shift_type(value):
    """
    Prekladá hodnoty z bunky Excelu na TypeShift ID.
    Prispôsob si IDs podľa tvojej databázy.
    """
    if not value:
        return None
        
    value = str(value).strip().upper() # Pre istotu pretypujeme na string
    
    if not value:
        return None
        
    # TOTO UPRAV PODĽA SVOJICH ID V DATABÁZE (tabuľka TypeShift)
    mapping = {
        '12': 1,    # Príklad: ID 1 je Denná 12ka
        'D': 1,     # Aj "D" znamená ID 1
        'N': 20,    # Príklad: ID 20 je Nočná
        '7': 5,     # 7h služba
        '9.5': 6,   # 9.5h služba
        'DO': 21,   # Dovolenka
        'P': 23,    # PN
        # Pridaj ďalšie skratky...
    }
    
    return mapping.get(value)